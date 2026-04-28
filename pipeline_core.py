"""
pipeline_core.py
All pipeline logic as a callable function — no global state, no hardcoded paths.
Called by app.py (Streamlit UI) and can also be called directly for testing.
"""

import os, re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ── Styles & formats ──────────────────────────────────────────────────────────
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
GREEN_FONT  = Font(color="008000", bold=True, size=11)
RED_FONT    = Font(color="FF0000", bold=True, size=11)
THIN        = Side(style='thin')
MEDIUM      = Side(style='medium')
DOUBLE      = Side(style='double')
CENTER_MID  = Alignment(horizontal='center', vertical='center')
EXCHANGE    = {'USD': 4.0, 'EUR': 4.7, 'CNY': 0.60, 'MYR': 1.0}
CHAIN_KW    = ['chain', 'tubing', 'hose']

FMT_ACCT = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
FMT_NUM  = '#,##0.00'
FMT_RM   = '_("RM"* #,##0.00_);_("RM"* (#,##0.00);_("RM"* "-"??_);_(@_)'

# ── Sheet configuration ───────────────────────────────────────────────────────
# (sname, afa, supp, ucur, urm, total, qty, note_col)
SHEET_CONFIG = {
    'AFA PAC Stock In.xlsx': [
        ('AFA PAC STOCK 1', 5, 4, 8, 9, 10, 7, 11),
        ('AFA PAC STOCK 2', 7, 6, 10, 11, 12, 9, 13),
    ],
    'AFA PAC Stock Out.xlsx': [
        ('STOCK 1', 6, 5, 9, 10, 11, 8, 12),
        ('STOCK 2', 6, 5, 9, 10, 11, 8, 12),
    ],
    'AFA Tech Stock In.xlsx': [
        ('AFA TRADING PARTS', 5, 4, 8, 9, 10, 7, 11),
        ('STOCK 1',           6, 5, 9, 10, 11, 8, 12),
        ('STOCK 2',           7, 6, 10, 11, 12, 9, 13),
        ('STOCK 3',           5, 4, 8, 9, 10, 7, 11),
        ('STOCK 3 - FMS',     5, 4, 8, 9, 10, 7, 11),
        ('STOCK 4',           7, 6, 10, 11, 12, 9, 13),
    ],
    'AFA Tech Stock Out.xlsx': [
        ('AFA TRAD',        6, 5, 9, 10, 11, 8, 12),
        ('Stock 1',         6, 5, 9, 10, 11, 8, 12),
        ('Stock 2',         6, 5, 9, 10, 11, 8, 12),
        ('Stock 3',         6, 5, 9, 10, 11, 8, 12),
        ('Stock 3-FMS2021', 6, 5, 9, 10, 11, 8, 12),
        ('Stock 4',         6, 5, 9, 10, 11, 8, 12),
    ],
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def to_float(v):
    try:    return float(v) if v is not None and not str(v).startswith('=') else None
    except: return None

def translate_formula(formula_str, new_row):
    if not formula_str or not str(formula_str).startswith('='): return None
    return re.sub(r'([A-Z]+)\d+', lambda m: m.group(1) + str(new_row), str(formula_str))

def border_all():  return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
def border_sum():  return Border(left=THIN, right=THIN, top=MEDIUM, bottom=DOUBLE)

def unbold_font(cell):
    f = cell.font
    try:    clr = f.color if f.color else None
    except: clr = None
    cell.font = Font(name=f.name, size=f.sz, bold=False, italic=f.italic,
                     underline=f.underline, strike=f.strike, color=clr)

def get_rate(note):
    if note is None: return None
    s = str(note).strip().upper()
    if s.startswith('RM') or s.startswith('MYR'): return 1.0
    for cur, r in EXCHANGE.items():
        if s.startswith(cur): return r
    return None

def find_last_and_sum(ws, afa_col, total_col, data_start=4):
    last_data_row = data_start - 1
    sum_row = None
    for r in range(data_start, ws.max_row + 2):
        v = ws.cell(r, total_col).value
        if v and str(v).upper().startswith('=SUM('):
            sum_row = r; break
        if ws.cell(r, afa_col).value is not None:
            last_data_row = r
    if sum_row is None:
        sum_row = last_data_row + 1
    return last_data_row, sum_row


# ── Main pipeline function ────────────────────────────────────────────────────
def run_pipeline(work_dir, month_label,
                 ref_pac_in=None, ref_pac_out=None,
                 ref_tech_in=None, ref_tech_out=None):
    """
    Run the full monthly stock pipeline.

    Parameters
    ----------
    work_dir     : str  — directory containing the 4 target xlsx + Master data.csv
    month_label  : str  — e.g. "April 2026"
    ref_pac_in   : str | None — path to PAC Stock Movement In reference file
    ref_pac_out  : str | None — path to PAC Stock Movement Out reference file
    ref_tech_in  : str | None — path to Tech Stock Movement In reference file
    ref_tech_out : str | None — path to Tech Stock Movement Out reference file

    Returns
    -------
    list[str]  — log lines (display in UI)
    """
    log = []
    def p(msg): log.append(msg)

    def fp(name): return os.path.join(work_dir, name)   # full path helper

    TASKS = [
        {'target': fp('AFA PAC Stock In.xlsx'),   'ref': ref_pac_in,
         'sheets': SHEET_CONFIG['AFA PAC Stock In.xlsx']},
        {'target': fp('AFA PAC Stock Out.xlsx'),  'ref': ref_pac_out,
         'sheets': SHEET_CONFIG['AFA PAC Stock Out.xlsx']},
        {'target': fp('AFA Tech Stock In.xlsx'),  'ref': ref_tech_in,
         'sheets': SHEET_CONFIG['AFA Tech Stock In.xlsx']},
        {'target': fp('AFA Tech Stock Out.xlsx'), 'ref': ref_tech_out,
         'sheets': SHEET_CONFIG['AFA Tech Stock Out.xlsx']},
    ]

    # ── fill_sheet (closure over log + master_lookup) ─────────────────────────
    def fill_sheet(ws_target, ws_ref_val, ws_ref_formula,
                   afa_col, supplier_col, unit_cur_col, unit_rm_col,
                   total_col, qty_col, note_col, data_start=4, sheet_label=''):
        qty_ltr   = get_column_letter(qty_col)
        urm_ltr   = get_column_letter(unit_rm_col)
        ucur_ltr  = get_column_letter(unit_cur_col)
        total_ltr = get_column_letter(total_col)
        desc_col  = afa_col + 1

        unmatched = []
        filled_ref = filled_mstr = 0
        last_data_row = data_start - 1

        for row in range(data_start, ws_target.max_row + 1):
            afa_val  = ws_target.cell(row, afa_col).value
            date_val = ws_target.cell(row, 2).value
            if afa_val is None and date_val is None:
                continue
            if afa_val is not None:
                last_data_row = row

            supp_val  = ws_target.cell(row, supplier_col).value
            ucur_val  = ws_target.cell(row, unit_cur_col).value
            urm_val   = ws_target.cell(row, unit_rm_col).value
            total_val = ws_target.cell(row, total_col).value

            ref_supp  = ws_ref_val.cell(row, supplier_col).value   if ws_ref_val   and row <= ws_ref_val.max_row   else None
            ref_ucur  = ws_ref_val.cell(row, unit_cur_col).value   if ws_ref_val   and row <= ws_ref_val.max_row   else None
            ref_urm   = ws_ref_val.cell(row, unit_rm_col).value    if ws_ref_val   and row <= ws_ref_val.max_row   else None
            ref_total = ws_ref_formula.cell(row, total_col).value  if ws_ref_formula and row <= ws_ref_formula.max_row else None

            if afa_val is not None:
                try:   code_key = str(int(float(str(afa_val)))) if str(afa_val).replace('.','').isdigit() else str(afa_val)
                except: code_key = str(afa_val)
            else:
                code_key = None

            # Fill Supplier
            if supp_val is None:
                if ref_supp:
                    ws_target.cell(row, supplier_col).value = ref_supp; filled_ref += 1
                elif code_key:
                    md = master_lookup.get(code_key)
                    if md and md['supplier']:
                        ws_target.cell(row, supplier_col).value = md['supplier']; filled_mstr += 1

            # Fill Unit Price Cur
            if ucur_val is None:
                if ref_ucur is not None:
                    ws_target.cell(row, unit_cur_col).value = ref_ucur
                elif code_key:
                    md = master_lookup.get(code_key)
                    if md and md['price_unit'] is not None:
                        ws_target.cell(row, unit_cur_col).value = md['price_unit']

            # Fill Unit Price RM
            if urm_val is None:
                if ref_urm is not None:
                    ws_target.cell(row, unit_rm_col).value = ref_urm
                elif code_key:
                    md = master_lookup.get(code_key)
                    if md and md['price_unit'] is not None:
                        rate = EXCHANGE.get(md.get('currency', 'MYR'), 1.0)
                        ws_target.cell(row, unit_rm_col).value = round(md['price_unit'] * rate, 4)

            # Fix total formula
            row_formula = f'={urm_ltr}{row}*{qty_ltr}{row}'
            total_str = str(total_val) if total_val is not None else ''
            is_premature_sum = total_str.upper().startswith('=SUM(')

            if is_premature_sum and (afa_val is not None or date_val is not None):
                ws_target.cell(row, total_col).value = row_formula
            elif ref_total and str(ref_total).startswith('=') \
                    and not str(ref_total).upper().startswith('=SUM('):
                new_formula = translate_formula(str(ref_total), row) or row_formula
                if total_val is None:
                    ws_target.cell(row, total_col).value = new_formula
                elif to_float(total_val) is not None:
                    urm_f = to_float(ws_target.cell(row, unit_rm_col).value)
                    qty_f = to_float(ws_target.cell(row, qty_col).value)
                    if urm_f is not None and qty_f is not None:
                        if abs(to_float(total_val) - round(urm_f * qty_f, 6)) > 0.001:
                            ws_target.cell(row, total_col).value = new_formula
                    else:
                        ws_target.cell(row, total_col).value = new_formula

            # Chain/tubing: sync ucur = urm
            desc = str(ws_target.cell(row, desc_col).value or '').lower()
            if any(kw in desc for kw in CHAIN_KW):
                urm_now  = ws_target.cell(row, unit_rm_col).value
                ucur_now = ws_target.cell(row, unit_cur_col).value
                try:
                    if urm_now is not None and ucur_now is not None \
                            and not str(urm_now).startswith('=') \
                            and abs(float(urm_now) - float(ucur_now)) > 0.0001:
                        ws_target.cell(row, unit_cur_col).value = urm_now
                except: pass

            # Missing conversion formula
            urm_raw = ws_target.cell(row, unit_rm_col).value
            if urm_raw is not None and not str(urm_raw).startswith('='):
                rate = get_rate(ws_target.cell(row, note_col).value)
                if rate is not None and rate != 1.0:
                    ucur_now = ws_target.cell(row, unit_cur_col).value
                    if ucur_now is not None:
                        try:
                            if abs(float(ucur_now) * rate - float(urm_raw)) <= max(0.02, abs(float(urm_raw)) * 0.01):
                                ws_target.cell(row, unit_rm_col).value = f'={ucur_ltr}{row}*{rate}'
                        except: pass

            # Special: AFA 10003157 — Sprocket chain priced RM26/ft, qty in mm
            # Note → "RM26/ft", ucur = 26, urm = =ucur/304.8 (ft→mm conversion)
            if code_key == '10003157':
                ws_target.cell(row, note_col).value      = 'RM26/ft'
                ws_target.cell(row, unit_cur_col).value  = 26.00
                ws_target.cell(row, unit_rm_col).value   = f'={ucur_ltr}{row}/304.8'

            # Unmatched check
            final_supp = ws_target.cell(row, supplier_col).value
            final_ucur = ws_target.cell(row, unit_cur_col).value
            if afa_val is not None and (final_supp is None or final_ucur is None):
                for c in range(1, min(ws_target.max_column + 1, total_col + 2)):
                    cell = ws_target.cell(row, c)
                    if cell.value is not None:
                        cell.fill = YELLOW_FILL
                unmatched.append((row, afa_val))

        # Place SUM formula — only when 2+ real data rows
        if last_data_row > data_start:
            sum_row = last_data_row + 1
            new_sum = f'=SUM({total_ltr}{data_start}:{total_ltr}{last_data_row})'
            for r in range(data_start, ws_target.max_row + 2):
                cell = ws_target.cell(r, total_col)
                if cell.value and str(cell.value).upper().startswith('=SUM('):
                    cell.value = None
            ws_target.cell(sum_row, total_col).value = new_sum

        p(f"    [{sheet_label}] ref={filled_ref} mstr={filled_mstr} unmatched={len(unmatched)}")
        return unmatched, last_data_row

    # ═══════════════════════════════════════════════════════════════════════════
    # STEP 1 — Load master data
    # ═══════════════════════════════════════════════════════════════════════════
    p("=" * 60)
    p(f"MONTHLY STOCK PIPELINE  —  {month_label}")
    p("=" * 60)
    p("\n[1] Loading Master data.csv ...")

    master_path = fp('Master data.csv.gz') if os.path.exists(fp('Master data.csv.gz')) else fp('Master data.csv')
    # ── Memory-efficient: stream CSV in chunks, keep only the latest row per code ──
    master_lookup = {}
    latest_date   = {}
    if not os.path.exists(master_path):
        p("    No Master data.csv uploaded — skipping master lookup (using ref files only).")
    else:
        cols = ['product_id/default_code', 'order_id/partner_id/name',
                'price_unit', 'order_id/currency_id/display_name', 'create_date']
        for chunk in pd.read_csv(master_path, encoding='utf-8', on_bad_lines='skip',
                                 dtype={'product_id/default_code': str},
                                 usecols=cols, chunksize=50_000):
            chunk['product_id/default_code'] = chunk['product_id/default_code'].astype(str).str.strip()
            chunk['create_date'] = pd.to_datetime(chunk['create_date'], errors='coerce')
            for _, row in chunk.iterrows():
                code = row['product_id/default_code']
                if not code or code == 'nan': continue
                d = row['create_date']
                prev = latest_date.get(code)
                if prev is not None and pd.notna(prev) and pd.notna(d) and d <= prev:
                    continue
                latest_date[code] = d
                master_lookup[code] = {
                    'supplier':   row['order_id/partner_id/name'] if pd.notna(row['order_id/partner_id/name']) else None,
                    'price_unit': float(row['price_unit']) if pd.notna(row['price_unit']) else None,
                    'currency':   str(row['order_id/currency_id/display_name']).strip()
                                  if pd.notna(row['order_id/currency_id/display_name']) else 'MYR',
                }
    del latest_date
    p(f"    {len(master_lookup)} unique AFA codes loaded.")

    # ═══════════════════════════════════════════════════════════════════════════
    # STEP 1b — Strip external links
    # ═══════════════════════════════════════════════════════════════════════════
    p("\n[1b] Removing external links from target files ...")
    for task in TASKS:
        fname = task['target']
        if not os.path.exists(fname): continue
        wb_v = load_workbook(fname, data_only=True)
        wb_f = load_workbook(fname)
        if not wb_f._external_links:
            wb_v.close(); wb_f.close()
            continue
        replaced = 0
        for sname in wb_f.sheetnames:
            ws_f = wb_f[sname]; ws_v = wb_v[sname]
            for r in range(1, ws_f.max_row + 1):
                for c in range(1, ws_f.max_column + 1):
                    v = ws_f.cell(r, c).value
                    if v and isinstance(v, str) and v.startswith('=') and '[' in v:
                        ws_f.cell(r, c).value = ws_v.cell(r, c).value
                        replaced += 1
        wb_f._external_links.clear()
        wb_v.close()
        tmp = fname + '.tmp.xlsx'
        wb_f.save(tmp); wb_f.close()
        os.replace(tmp, fname)
        p(f"    {os.path.basename(fname)}: cleared external links, replaced {replaced} formula(s)")

    # ═══════════════════════════════════════════════════════════════════════════
    # STEP 2 — Fill data from reference + master
    # ═══════════════════════════════════════════════════════════════════════════
    p("\n[2] Filling data from reference files ...")
    NOTE_COL = 15

    for task in TASKS:
        target_file = task['target']
        ref_file    = task['ref']
        if not os.path.exists(target_file): continue

        p(f"\n  {os.path.basename(target_file)}")
        wb_target = load_workbook(target_file)

        if ref_file and os.path.exists(ref_file):
            wb_ref_val     = load_workbook(ref_file, data_only=True)
            wb_ref_formula = load_workbook(ref_file)
        else:
            if ref_file:
                p(f"  WARNING: Reference not found — filling from master data only")
            wb_ref_val = wb_ref_formula = None

        for row in task['sheets']:
            sname, afa, supp, ucur, urm, total, qty, note = row
            if sname not in wb_target.sheetnames:
                p(f"    SKIP {sname}"); continue

            ws_rv = wb_ref_val[sname]     if wb_ref_val     else None
            ws_rf = wb_ref_formula[sname] if wb_ref_formula else None

            unmatched, _ = fill_sheet(
                wb_target[sname], ws_rv, ws_rf,
                afa, supp, ucur, urm, total, qty, note, sheet_label=sname
            )

            # Status note at O1 — carve out col 15, re-merge left/right
            ws_note = wb_target[sname]
            for mr in list(ws_note.merged_cells.ranges):
                if mr.min_row <= 1 <= mr.max_row and mr.min_col <= NOTE_COL <= mr.max_col:
                    lo, hi, top, bot = mr.min_col, mr.max_col, mr.min_row, mr.max_row
                    ws_note.unmerge_cells(str(mr))
                    if lo < NOTE_COL:
                        ws_note.merge_cells(start_row=top, start_column=lo,
                                            end_row=bot, end_column=NOTE_COL - 1)
                    if hi > NOTE_COL:
                        ws_note.merge_cells(start_row=top, start_column=NOTE_COL + 1,
                                            end_row=bot, end_column=hi)
            ws_note.cell(1, 9).value     = None
            ws_note.cell(1, 9).font      = Font()
            ws_note.cell(1, 9).alignment = Alignment()

            note_text = "All items successfully matched and filled from OneDrive reference / master data." \
                        if not unmatched else \
                        f"WARNING: {len(unmatched)} unmatched — AFA codes: {', '.join(str(a) for _, a in unmatched)}"
            cell = ws_note.cell(1, NOTE_COL)
            cell.value = note_text
            cell.font  = GREEN_FONT if not unmatched else RED_FONT
            cell.alignment = Alignment(wrap_text=True, vertical='top')

        wb_target.save(target_file)
        if wb_ref_val:     wb_ref_val.close()
        if wb_ref_formula: wb_ref_formula.close()

    # ═══════════════════════════════════════════════════════════════════════════
    # STEP 3 — Formatting
    # ═══════════════════════════════════════════════════════════════════════════
    p("\n[3] Applying formatting ...")

    for task in TASKS:
        fname = task['target']
        if not os.path.exists(fname): continue
        wb = load_workbook(fname)

        for row in task['sheets']:
            sname, afa_col, supp_col, ucur_col, urm_col, total_col, qty_col, note_col = row
            if sname not in wb.sheetnames: continue
            ws = wb[sname]

            max_col = 1
            for c in range(ws.max_column, 0, -1):
                if ws.cell(3, c).value is not None: max_col = c; break

            last_data_row, sum_row = find_last_and_sum(ws, afa_col, total_col)
            total_ltr = get_column_letter(total_col)
            urm_ltr   = get_column_letter(urm_col)
            qty_ltr   = get_column_letter(qty_col)

            for r in range(4, sum_row + 1):
                is_sum = (r == sum_row)
                for c in range(1, max_col + 1):
                    cell = ws.cell(r, c)
                    cell.border = border_sum() if (is_sum and c == total_col) else border_all()
                    if is_sum and c == total_col:
                        cell.font = Font(bold=True)
                    elif not is_sum:
                        unbold_font(cell)
                ws.cell(r, qty_col).number_format   = FMT_ACCT
                ws.cell(r, ucur_col).number_format  = FMT_ACCT
                ws.cell(r, urm_col).number_format   = FMT_ACCT
                ws.cell(r, total_col).number_format = FMT_RM if is_sum else FMT_NUM
                ws.cell(r, total_col).alignment     = CENTER_MID
                if not is_sum and ws.cell(r, afa_col).value is not None:
                    ws.cell(r, total_col).value = f'={urm_ltr}{r}*{qty_ltr}{r}'
                ws.cell(r, qty_col).alignment = Alignment(
                    wrap_text=False,
                    horizontal=ws.cell(r, qty_col).alignment.horizontal,
                    vertical=ws.cell(r, qty_col).alignment.vertical)

            ws.cell(3, total_col).value = 'Inventory Cost (RM)'
            ws.cell(1, 9).value     = None
            ws.cell(1, 9).font      = Font()
            ws.cell(1, 9).alignment = Alignment()

            seq = 0
            for r in range(4, last_data_row + 1):
                if ws.cell(r, afa_col).value is not None:
                    seq += 1
                    ws.cell(r, 1).value = seq
                    ws.cell(r, 1).alignment = Alignment(horizontal='center', vertical='center')

            # ── Final override: AFA 10003157 — RM26/ft sprocket chain (qty in mm) ──
            ucur_ltr_local = get_column_letter(ucur_col)
            for r in range(4, sum_row):
                v = ws.cell(r, afa_col).value
                if v is None: continue
                try:
                    code = str(int(float(str(v)))) if str(v).replace('.','').isdigit() else str(v).strip()
                except:
                    code = str(v).strip()
                if code == '10003157':
                    ws.cell(r, note_col).value     = 'RM26/ft'
                    ws.cell(r, ucur_col).value     = 26.00
                    ws.cell(r, urm_col).value      = f'={ucur_ltr_local}{r}/304.8'
                    ws.cell(r, ucur_col).number_format = FMT_ACCT
                    ws.cell(r, urm_col).number_format  = FMT_ACCT

        # AFA Tech Stock Out: restore E1:G1 merge, move A1 content, clear A1
        if 'Tech Stock Out' in fname:
            for row in task['sheets']:
                sname = row[0]
                if sname not in wb.sheetnames: continue
                ws = wb[sname]
                a1 = ws.cell(1, 1)
                a1_value = a1.value
                for mr in list(ws.merged_cells.ranges):
                    if mr.min_row <= 1 <= mr.max_row and mr.min_col <= 7 and mr.max_col >= 5:
                        ws.unmerge_cells(str(mr))
                ws.merge_cells('E1:G1')
                e1 = ws.cell(1, 5)
                e1.value     = a1_value if a1_value else f'{sname} Report'
                e1.font      = Font(bold=True, size=12)
                e1.alignment = Alignment(horizontal='center', vertical='center')
                a1.value     = None
                a1.font      = Font()
                a1.alignment = Alignment()

        wb.save(fname)
        p(f"    Saved: {os.path.basename(fname)}")

    p("\n" + "=" * 60)
    p("PIPELINE COMPLETE")
    p("=" * 60)
    return log
